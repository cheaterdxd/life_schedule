import tkinter as tk
from tkinter import simpledialog, messagebox, PhotoImage
import time, threading
import os
import sys # Dùng để lấy đường dẫn nếu đóng gói thành exe
import datetime # Thêm thư viện datetime để xử lý ngày tháng
import json
# Thư viện bên ngoài (cần cài đặt: pip install openpyxl playsound3)
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    messagebox.showerror("Lỗi thiếu thư viện", "Vui lòng cài đặt thư viện openpyxl: pip install openpyxl")
    exit()
try:
    from playsound3 import playsound
except ImportError:
    messagebox.showerror("Lỗi thiếu thư viện", "Vui lòng cài đặt thư viện playsound3: pip install playsound3")
    exit()

# --- Cấu hình ---
DEFAULT_INTERVAL_MINUTES = 15
EXCEL_FILE = "activity_log.xlsx"
SOUND_FILE = "clockbeep.wav"     # Âm thanh nhắc nhở hoạt động (cũ)
ALERT_SOUND = "alert.wav"        # Âm thanh cảnh báo (mới cho nhiệm vụ)
TODO_FILE_PREFIX = "todolist_"   # Tiền tố file nhiệm vụ

# --- Hàm tiện ích tìm đường dẫn (hữu ích khi đóng gói bằng PyInstaller) ---
def resource_path(relative_path):
    """ Lấy đường dẫn tuyệt đối đến tài nguyên, hoạt động cho cả dev và PyInstaller """
    try:
        # PyInstaller tạo thư mục tạm _MEIPASS và lưu trữ tài nguyên ở đó
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- Lớp ứng dụng chính ---
class ReminderApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Trình Nhắc Nhở Hoạt Động & Nhiệm vụ") # Cập nhật tiêu đề
        # self.root.geometry("300x180") # Có thể cần tăng kích thước một chút
        self.root.resizable(False, False)

        # --- Biến trạng thái nhắc nhở hoạt động (cũ) ---
        self.reminder_interval_ms = DEFAULT_INTERVAL_MINUTES * 60 * 1000
        self.next_reminder_id = None
        self.sleep_until_time = 0
        self.countdown_id = None
        self.is_running = True

        # --- Biến trạng thái nhiệm vụ (mới) ---
        self.tasks = [] # Danh sách lưu nhiệm vụ: [{'name': 'Task Name', 'time': 'HH:MM', 'completed': False, 'alarm_id': None}]
        self.task_window = None # Theo dõi cửa sổ xem nhiệm vụ
        self.task_creation_win = None # Theo dõi cửa sổ tạo nhiệm vụ
        self._temp_tasks_to_add = [] # Danh sách tạm nhiệm vụ khi đang tạo

        # --- Âm thanh và Icon ---
        self.sound_path = resource_path(SOUND_FILE)
        self.alert_sound_path = resource_path(ALERT_SOUND) # Đường dẫn file âm thanh cảnh báo
        # Thử tải icon (bỏ qua nếu không có file icon.png)
        try:
            icon_path = resource_path("icon.png")
            if os.path.exists(icon_path):
                 self.root.iconphoto(True, PhotoImage(file=icon_path))
        except Exception as e:
            print(f"Không thể tải icon: {e}")

        # --- Giao diện ---
        # Label hiển thị trạng thái chung (Đang chạy/Đang ngủ)
        self.status_label = tk.Label(root_window, text="Đang chạy...", font=("Helvetica", 10), pady=5)
        self.status_label.pack(fill=tk.X, padx=10)

        # Label hiển thị thời gian kiểm tra hoạt động cuối
        self.last_check_label = tk.Label(root_window, text="Chưa có lần kiểm tra hoạt động nào.", font=("Helvetica", 9), fg="gray", pady=5)
        self.last_check_label.pack(fill=tk.X, padx=10)

        button_frame = tk.Frame(root_window)
        button_frame.pack(pady=10)

        # Nút Đi ngủ (Sleep Button)
        self.sleep_button = tk.Button(button_frame,
                                       text="😴 Đi ngủ",
                                       command=self.go_to_sleep,
                                       width=10,
                                       bg="lightblue",
                                       fg="black",
                                       activebackground="dodgerblue",
                                       activeforeground="white",
                                       relief="flat")
        self.sleep_button.grid(row=0, column=0, padx=5)

        # Nút Cài đặt (Settings Button)
        self.settings_button = tk.Button(button_frame,
                                         text="⚙️ Cài đặt",
                                         command=self.change_interval,
                                         width=10,
                                         bg="lightgray",
                                         fg="black",
                                         activebackground="gray",
                                         activeforeground="white",
                                         relief="flat")
        self.settings_button.grid(row=0, column=1, padx=5)

        # Nút Dừng (Stop Button)
        self.stop_button = tk.Button(button_frame,
                                     text="⏹️ Dừng",
                                     command=self.stop_program,
                                     width=10,
                                     bg="red",
                                     fg="white",
                                     activebackground="darkred",
                                     activeforeground="white",
                                     relief="flat")
        self.stop_button.grid(row=0, column=2, padx=5)

        # --- Thêm Nút Nhiệm vụ (mới) ---
        self.tasks_button = tk.Button(button_frame,
                                      text="📝 Nhiệm vụ",
                                      command=self.show_task_window, # Mở cửa sổ quản lý nhiệm vụ
                                      width=10,
                                      bg="orange",
                                      fg="black",
                                      activebackground="darkorange",
                                      activeforeground="white",
                                      relief="flat")
        self.tasks_button.grid(row=0, column=3, padx=5) # Đặt nút này vào cột thứ 3

        # Đảm bảo file Excel tồn tại với tiêu đề
        self.ensure_excel_file()

        # --- Khởi tạo và kiểm tra/tải nhiệm vụ khi ứng dụng khởi động ---
        self.initialize_tasks()

        # Bắt đầu vòng lặp nhắc nhở hoạt động (chỉ khi không đang ngủ)
        # Vòng lặp này sẽ được gọi ở cuối initialize_tasks()
        # self.schedule_next_reminder()


    # --- Phương thức mới liên quan đến Nhiệm vụ ---

    def initialize_tasks(self):
        """Kiểm tra file todolist của hôm nay, tải nhiệm vụ nếu có."""
        print("Đang kiểm tra và tải nhiệm vụ...")
        tasks_exist = self.check_and_load_todolist() # Trả về True nếu file tồn tại và đọc được

        if not tasks_exist or not self.tasks: # Nếu file không tồn tại HOẶC file rỗng sau khi tải
            print("Không tìm thấy file nhiệm vụ hoặc file trống. Phát cảnh báo và yêu cầu tạo.")
            self.play_alert_sound()
            messagebox.showinfo("Thông báo Nhiệm vụ",
                                f"Không tìm thấy file nhiệm vụ cho hôm nay ({datetime.date.today().strftime('%d/%m/%Y')}). Vui lòng tạo nhiệm vụ mới.",
                                parent=self.root)
            # Có thể thêm lựa chọn tự động mở cửa sổ tạo nhiệm vụ ở đây

        # --- Lên lịch báo thức cho các nhiệm vụ chưa hoàn thành (sẽ triển khai sau) ---
        # self.schedule_task_alarms() # Hàm này sẽ được gọi ở bước sau

        # Sau khi xử lý nhiệm vụ ban đầu, bắt đầu vòng lặp nhắc nhở hoạt động cũ
        # Chỉ bắt đầu nếu không đang trong chế độ ngủ từ lần chạy trước (trạng thái ngủ không được lưu)
        # Hiện tại chỉ cần gọi như cũ là đủ vì schedule_next_reminder đã kiểm tra sleep_until_time
        self.schedule_next_reminder()

    def ensure_excel_file(self):
        """Kiểm tra và tạo file Excel nếu chưa có."""
        excel_path = resource_path(EXCEL_FILE)
        if not os.path.exists(excel_path):
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(["Thời gian", "Hoạt động"])
                wb.save(excel_path)
                print(f"Đã tạo file log: {excel_path}")
            except Exception as e:
                messagebox.showerror("Lỗi File Excel", f"Không thể tạo file {EXCEL_FILE}:\n{e}", parent=self.root)
                self.stop_program() # Dừng nếu không tạo được file log

    def get_today_todo_filepath(self):
        """Trả về đường dẫn đầy đủ đến file nhiệm vụ của ngày hôm nay (YYYYMMDD.json)."""
        today_str = datetime.date.today().strftime("%Y%m%d")
        # --- Đổi đuôi file nhiệm vụ sang .json ---
        filename = f"{TODO_FILE_PREFIX}{today_str}.json"
        # --------------------------------------
        return resource_path(filename)

    def check_and_load_todolist(self):
        """Kiểm tra file nhiệm vụ JSON của hôm nay, tải nội dung nếu có.
           Trả về True nếu file tồn tại và không có lỗi đọc, False nếu không tồn tại hoặc lỗi."""
        filepath = self.get_today_todo_filepath() # Sử dụng đường dẫn file .json
        self.tasks = [] # Xóa danh sách nhiệm vụ hiện tại trước khi tải

        if os.path.exists(filepath):
            print(f"Tìm thấy file nhiệm vụ: {filepath}")
            try:
                # --- Đọc file JSON ---
                with open(filepath, 'r', encoding='utf-8') as f:
                    # Kiểm tra file không rỗng trước khi load
                    content = f.read()
                    if content:
                         # Parse nội dung JSON thành danh sách các dictionary Python
                         self.tasks = json.loads(content)
                    else:
                         self.tasks = [] # File rỗng, danh sách nhiệm vụ trống

                # --- Xử lý tương thích ngược và thêm các khóa mặc định nếu file cũ thiếu ---
                # Điều này giúp ứng dụng hoạt động với file JSON cũ hơn (nếu có)
                for task in self.tasks:
                    # alarm_id không được lưu trong file, cần khởi tạo lại mỗi lần tải
                    task['alarm_id'] = None
                    # Thêm các khóa mới nếu file JSON cũ không có chúng (đảm bảo cấu trúc đồng nhất)
                    if 'completed' not in task:
                        task['completed'] = False # Mặc định là False nếu không có
                    if 'completion_time' not in task:
                         task['completion_time'] = None # Mặc định là None nếu không có
                    if 'labels' not in task:
                         task['labels'] = [] # Mặc định là danh sách rỗng nếu không có
                    if 'scheduled_time' not in task:
                         # Mặc định scheduled_time lấy từ 'time' nếu không có khóa scheduled_time
                         task['scheduled_time'] = task.get('time')
                    # Đảm bảo có ít nhất khóa 'name' và 'time' để nhiệm vụ hợp lệ
                    if 'name' not in task or 'time' not in task:
                         print(f"Cảnh báo: Nhiệm vụ có cấu trúc không đầy đủ bị bỏ qua khi tải: {task}")
                         # Sử dụng list slicing hoặc list comprehension để xóa an toàn khi lặp
                         continue # Bỏ qua nhiệm vụ không hợp lệ trong vòng lặp for

                # Sau vòng lặp, lọc lại danh sách self.tasks để bỏ các nhiệm vụ không hợp lệ
                self.tasks = [task for task in self.tasks if 'name' in task and 'time' in task]


                print(f"Đã tải {len(self.tasks)} nhiệm vụ từ file JSON.")
                return True
            except json.JSONDecodeError as e:
                 print(f"Lỗi giải mã JSON file nhiệm vụ {filepath}: {e}")
                 messagebox.showerror("Lỗi File Nhiệm vụ", f"Không thể đọc file nhiệm vụ {filepath} (lỗi JSON):\n{e}", parent=self.root)
                 # Tùy chọn: Đổi tên file bị lỗi để không đọc lại lần sau và cho người dùng kiểm tra
                 # os.rename(filepath, filepath + ".backup." + time.strftime("%Y%m%d%H%M%S"))
                 return False
            except Exception as e:
                print(f"Lỗi khi đọc file nhiệm vụ {filepath}:\n{e}")
                messagebox.showerror("Lỗi File Nhiệm vụ", f"Không thể đọc file nhiệm vụ {filepath}:\n{e}", parent=self.root)
                return False
        else:
            print(f"Không tìm thấy file nhiệm vụ: {filepath}")
            return False # File không tồn tại

    def save_todolist(self):
        """Lưu danh sách nhiệm vụ hiện tại (từ self.tasks) vào file JSON của ngày hôm nay."""
        filepath = self.get_today_todo_filepath() # Sử dụng đường dẫn file .json

        # Chuẩn bị dữ liệu để lưu: loại bỏ các khóa không cần thiết (như alarm_id)
        tasks_to_save = []
        for task in self.tasks:
            task_copy = task.copy() # Tạo bản sao để không ảnh hưởng đến dictionary gốc
            # Kiểm tra sự tồn tại của khóa 'alarm_id' trước khi cố gắng xóa
            if 'alarm_id' in task_copy:
                del task_copy['alarm_id'] # ID báo thức chỉ tồn tại khi ứng dụng chạy, không lưu vào file

            # Thêm bản sao đã loại bỏ 'alarm_id' vào danh sách sẽ ghi ra file
            tasks_to_save.append(task_copy)

        try:
            # --- Ghi file JSON ---
            # use ensure_ascii=False để Python ghi ký tự tiếng Việt mà không mã hóa (ví dụ: \u1ef9)
            # use indent=4 để file JSON được định dạng đẹp, dễ đọc hơn khi mở bằng trình soạn thảo văn bản
            with open(filepath, 'w', encoding='utf-8') as f:
                 json.dump(tasks_to_save, f, ensure_ascii=False, indent=4)
            print(f"Đã lưu {len(tasks_to_save)} nhiệm vụ vào file JSON: {filepath}")
        except Exception as e:
            print(f"Lỗi khi ghi file nhiệm vụ {filepath}: {e}")
            messagebox.showerror("Lỗi Lưu Nhiệm vụ", f"Không thể ghi file nhiệm vụ {filepath}:\n{e}", parent=self.root)

    def play_alert_sound(self):
        """Phát âm thanh cảnh báo (khi không tìm thấy file nhiệm vụ)."""
        try:
            if os.path.exists(self.alert_sound_path):
                 # Chạy playsound trong thread riêng để tránh treo GUI
                threading.Thread(target=playsound, args=(self.alert_sound_path,), daemon=True).start()
            else:
                print(f"Không tìm thấy file âm thanh cảnh báo: {self.alert_sound_path}")
                # Tùy chọn: hiển thị messagebox nếu không tìm thấy file âm thanh
                # messagebox.showwarning("Lỗi Âm thanh", f"Không tìm thấy file âm thanh cảnh báo:\n{os.path.basename(ALERT_SOUND)}", parent=self.root)
        except Exception as e:
            print(f"Lỗi khi phát âm thanh cảnh báo '{self.alert_sound_path}': {e}")
            # Tránh messagebox ở đây để không làm phiền liên tục

    def show_task_window(self):
        """Hiển thị cửa sổ quản lý/xem nhiệm vụ của hôm nay."""
        # Tránh mở nhiều cửa sổ cùng loại
        if self.task_window is None or not self.task_window.winfo_exists():
            self.task_window = tk.Toplevel(self.root)
            self.task_window.title(f"Nhiệm vụ hôm nay ({datetime.date.today().strftime('%d/%m/%Y')})")
            # self.task_window.geometry("400x300") # Kích thước tùy chọn
            self.task_window.transient(self.root) # Làm cửa sổ con của main window
            self.task_window.grab_set() # Chặn tương tác với cửa sổ khác chính

            # --- Nội dung cửa sổ Nhiệm vụ ---
            label = tk.Label(self.task_window, text="Danh sách nhiệm vụ của bạn:")
            label.pack(pady=5)

            # Nút mở cửa sổ tạo nhiệm vụ mới
            add_button = tk.Button(self.task_window, text="➕ Thêm Nhiệm vụ mới", command=self.show_task_creation_window)
            add_button.pack(pady=5)

            # Placeholder: Khu vực hiển thị danh sách nhiệm vụ
            # Việc hiển thị từng nhiệm vụ với trạng thái và nút hoàn thành sẽ cần triển khai sau.
            # Tạm thời, chỉ hiển thị thông báo nếu không có nhiệm vụ.
            if not self.tasks:
                 no_tasks_label = tk.Label(self.task_window, text="Chưa có nhiệm vụ nào được thêm cho hôm nay.", fg="gray")
                 no_tasks_label.pack(pady=10)
            else:
                 # --- HIỂN THỊ DANH SÁCH NHIỆM VỤ (CẦN TRIỂN KHAI THÊM) ---
                 # Vòng lặp dưới đây chỉ là ví dụ đơn giản hiển thị tên và thời gian
                 # Bạn sẽ cần tạo các widget (Label, Checkbutton/Button) cho từng nhiệm vụ
                 # và thêm cơ chế đánh dấu hoàn thành.
                 for task in self.tasks:
                     task_text = f"{task['time']} - {task['name']}"
                     # Cần hiển thị trạng thái completed[''] sau
                     task_label = tk.Label(self.task_window, text=task_text)
                     task_label.pack(anchor='w', padx=10) # Căn trái


                 # --- KẾT THÚC HIỂN THỊ DANH SÁCH NHIỆM VỤ ---


            # Nút đóng cửa sổ nhiệm vụ
            close_button = tk.Button(self.task_window, text="Đóng", command=self.task_window.destroy)
            close_button.pack(pady=10)

            # Xử lý khi đóng cửa sổ bằng nút 'X'
            self.task_window.protocol("WM_DELETE_WINDOW", self.task_window.destroy)


        self.task_window.lift() # Đưa cửa sổ lên trên cùng


    def show_task_creation_window(self):
        """Hiển thị cửa sổ cho phép người dùng nhập nhiệm vụ mới."""
        # Tránh mở nhiều cửa sổ tạo nhiệm vụ
        if self.task_creation_win is None or not self.task_creation_win.winfo_exists():
            self.task_creation_win = tk.Toplevel(self.root)
            self.task_creation_win.title("Tạo Nhiệm vụ mới")
            self.task_creation_win.geometry("350x220") # Có thể cần điều chỉnh kích thước
            self.task_creation_win.transient(self.root) # Làm cửa sổ con
            self.task_creation_win.grab_set() # Chặn tương tác với cửa sổ khác

            # Danh sách tạm thời để giữ các nhiệm vụ được thêm trong phiên làm việc của cửa sổ này
            # Chỉ thêm vào self.tasks và lưu file khi nhấn nút "Lưu và Đóng"
            self._temp_tasks_to_add = []

            tk.Label(self.task_creation_win, text="Tên Nhiệm vụ:").pack(pady=(10,0))
            task_name_entry = tk.Entry(self.task_creation_win, width=40)
            task_name_entry.pack(pady=(0,5))
            task_name_entry.focus_set() # Focus vào ô nhập tên

            tk.Label(self.task_creation_win, text="Thời gian (HH:MM - 24h):").pack(pady=(10,0))
            task_time_entry = tk.Entry(self.task_creation_win, width=10)
            task_time_entry.pack(pady=(0,5))

            # Frame cho các nút hành động
            button_frame = tk.Frame(self.task_creation_win)
            button_frame.pack(pady=10)

            # Nút "Thêm vào danh sách tạm thời"
            add_button = tk.Button(button_frame, text="➕ Thêm vào danh sách tạm",
                                   command=lambda: self.add_task_from_entries(task_name_entry, task_time_entry))
            add_button.grid(row=0, column=0, padx=5)

            # Nút "Lưu và Đóng" (Lưu các nhiệm vụ tạm vào danh sách chính và file)
            save_button = tk.Button(button_frame, text="💾 Lưu tất cả và Đóng",
                                    command=self.save_and_close_task_creation)
            save_button.grid(row=0, column=1, padx=5)

            # Nút "Hủy" (Đóng cửa sổ mà không lưu các nhiệm vụ tạm)
            cancel_button = tk.Button(button_frame, text="❌ Hủy",
                                      command=self.task_creation_win.destroy)
            cancel_button.grid(row=0, column=2, padx=5)

            # Có thể thêm một khu vực hiển thị các nhiệm vụ đã thêm tạm ở đây nếu muốn

            # Xử lý khi đóng cửa sổ bằng nút 'X' (tương tự như nhấn Hủy)
            self.task_creation_win.protocol("WM_DELETE_WINDOW", self.task_creation_win.destroy)


        self.task_creation_win.lift() # Đưa cửa sổ tạo nhiệm vụ lên trên cùng


    def add_task_from_entries(self, task_name_entry, task_time_entry):
        """Lấy dữ liệu từ Entry trong cửa sổ tạo nhiệm vụ, validate và thêm vào danh sách tạm thời."""
        name = task_name_entry.get().strip()
        time_str = task_time_entry.get().strip()

        if not name:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập tên nhiệm vụ.", parent=self.task_creation_win)
            return
        if not time_str:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập thời gian nhiệm vụ (HH:MM).", parent=self.task_creation_win)
            return

        # Xác thực định dạng thời gian cơ bản (HH:MM)
        try:
            # Kiểm tra định dạng HH:MM và giá trị hợp lệ
            hours, minutes = map(int, time_str.split(':'))
            if not (0 <= hours <= 23 and 0 <= minutes <= 59):
                 raise ValueError("Giờ hoặc phút không hợp lệ")
            # Chỉ cần kiểm tra định dạng, không cần đối tượng time_obj ở đây
            datetime.datetime.strptime(time_str, "%H:%M")
        except ValueError:
            messagebox.showwarning("Sai định dạng", "Thời gian không hợp lệ. Vui lòng nhập theo định dạng HH:MM (ví dụ: 09:00, 14:30).", parent=self.task_creation_win)
            return

        # Thêm nhiệm vụ vào danh sách tạm thời
        self._temp_tasks_to_add.append({'name': name, 'time': time_str, 'completed': False, 'alarm_id': None})
        print(f"Đã thêm tạm nhiệm vụ: {name} lúc {time_str}")

        # Thông báo đã thêm và làm sạch các ô nhập liệu
        messagebox.showinfo("Thành công", f"Đã thêm tạm nhiệm vụ '{name}'. Tiếp tục nhập nhiệm vụ khác hoặc nhấn 'Lưu'.", parent=self.task_creation_win)
        task_name_entry.delete(0, tk.END)
        task_time_entry.delete(0, tk.END) # Xóa cả thời gian để người dùng nhập mới cho nhiệm vụ tiếp theo
        task_name_entry.focus_set() # Trả focus về ô tên nhiệm vụ

        # Tùy chọn: Cập nhật khu vực xem trước nhiệm vụ tạm trong cửa sổ tạo

    def save_and_close_task_creation(self):
        """Lưu tất cả các nhiệm vụ từ danh sách tạm thời vào danh sách chính và file, sau đó đóng cửa sổ tạo."""
        if not self._temp_tasks_to_add:
            messagebox.showinfo("Thông báo", "Chưa có nhiệm vụ mới nào được thêm để lưu.", parent=self.task_creation_win)
            self.task_creation_win.destroy()
            return

        # Thêm các nhiệm vụ từ danh sách tạm vào danh sách chính
        # Lưu ý: Hiện tại các nhiệm vụ cũ trong self.tasks sẽ bị giữ lại.
        # Nếu bạn muốn chỉ lưu nhiệm vụ MỚI từ cửa sổ này, bạn có thể clear self.tasks trước
        # hoặc hỏi người dùng có muốn ghi đè/thêm vào.
        # Cách hiện tại là THÊM VÀO danh sách nhiệm vụ hiện có.
        self.tasks.extend(self._temp_tasks_to_add)
        self._temp_tasks_to_add = [] # Xóa danh sách tạm sau khi đã thêm vào chính

        # Lưu danh sách nhiệm vụ chính (bao gồm cả cũ và mới được thêm tạm) vào file
        self.save_todolist()

        # --- Cần triển khai sau: Lên lịch lại báo thức cho TẤT CẢ nhiệm vụ chưa hoàn thành ---
        # (bao gồm cả nhiệm vụ cũ và mới)
        # self.cancel_all_task_alarms() # Hủy hết báo thức cũ (nếu có)
        # self.schedule_task_alarms() # Lên lịch lại

        # --- Cần triển khai sau: Cập nhật hiển thị trong cửa sổ xem nhiệm vụ nếu nó đang mở ---
        # self.update_task_display()
        if self.task_window and self.task_window.winfo_exists():
             # Tùy chọn: đóng cửa sổ xem nhiệm vụ cũ và mở lại để cập nhật dễ hơn
             # Hoặc viết hàm update_task_display để làm mới nội dung
             self.task_window.destroy()
             self.show_task_window()


        messagebox.showinfo("Lưu thành công", f"Đã lưu tổng cộng {len(self.tasks)} nhiệm vụ vào file.", parent=self.task_creation_win)
        self.task_creation_win.destroy() # Đóng cửa sổ tạo nhiệm vụ


    # --- Phương thức liên quan đến nhắc nhở hoạt động (cũ) ---

    # Các phương thức save_log, play_notification_sound, schedule_next_reminder,
    # ask_question, go_to_sleep, update_countdown, change_interval, stop_program
    # giữ nguyên hoặc chỉ sửa đổi nhỏ (như cập nhật last_check_label trong save_log)

    def save_log(self, answer):
        """Lưu câu trả lời vào file Excel và cập nhật Label thời gian."""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        excel_path = resource_path(EXCEL_FILE)
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            ws.append([timestamp, answer])
            wb.save(excel_path)
            # --- Cập nhật Label thời gian kiểm tra cuối ---
            self.last_check_label.config(text=f"Lần cuối kiểm tra hoạt động lúc: {timestamp}", fg="black") # Đổi lại màu chữ cho rõ hơn
            # ---------------------------------------------
        except Exception as e:
             messagebox.showwarning("Lỗi Lưu Log", f"Không thể ghi vào file {EXCEL_FILE}:\n{e}\n\nLog sẽ chỉ được lưu tạm vào bộ nhớ.", parent=self.root)
             # Có thể thêm cơ chế lưu tạm vào biến hoặc file text dự phòng ở đây

    def play_notification_sound(self):
        """Phát âm thanh thông báo nhắc nhở hoạt động."""
        try:
            # Chạy playsound trong thread riêng để tránh treo GUI nếu file lớn/lỗi
            threading.Thread(target=playsound, args=(self.sound_path,), daemon=True).start()
        except Exception as e:
            print(f"Lỗi khi phát âm thanh nhắc nhở '{self.sound_path}': {e}")
            # Không cần messagebox ở đây để tránh làm phiền liên tục


    def schedule_next_reminder(self):
        """Lên lịch cho lần nhắc nhở hoạt động tiếp theo."""
        # Hủy lịch trình cũ nếu có
        if self.next_reminder_id:
            self.root.after_cancel(self.next_reminder_id)
            self.next_reminder_id = None # Đặt lại về None sau khi hủy

        if self.is_running and time.time() >= self.sleep_until_time:
            print(f"Lên lịch nhắc nhở hoạt động sau {self.reminder_interval_ms // 1000} giây.")
            # Tính toán thời gian chờ dựa trên interval
            delay_ms = self.reminder_interval_ms
            self.next_reminder_id = self.root.after(delay_ms, self.ask_question)
            # Cập nhật status_label nếu không đang trong chế độ ngủ
            if self.countdown_id is None:
                 self.status_label.config(text=f"Đang chạy... Nhắc nhở sau {delay_ms // 60000} phút")
        elif self.is_running and time.time() < self.sleep_until_time:
            # Nếu đang ngủ, lên lịch để kiểm tra lại khi hết giờ ngủ
             wake_up_delay = int((self.sleep_until_time - time.time()) * 1000)
             # Đảm bảo delay không âm
             if wake_up_delay < 0:
                 wake_up_delay = 0
             print(f"Đang ngủ, kiểm tra lại sau {wake_up_delay / 1000:.0f} giây.")
             # Lên lịch kiểm tra lại. update_countdown sẽ cập nhật status_label.
             # Thêm 0.5s để chắc chắn đã qua thời điểm sleep_until_time một chút
             self.next_reminder_id = self.root.after(wake_up_delay + 500, self.schedule_next_reminder)
             # status_label đã được update_countdown xử lý khi đang ngủ

        # Nếu self.is_running là False, sẽ không lên lịch nhắc nhở mới

    def ask_question(self):
        """Hỏi người dùng về hoạt động, yêu cầu nhập, và lên lịch lại."""
        # Kiểm tra lại trạng thái trước khi hiển thị cửa sổ
        # Đảm bảo cửa sổ không hiện lên nếu ứng dụng đã dừng hoặc đang ngủ
        if not self.is_running or time.time() < self.sleep_until_time:
            print("Bỏ qua cửa sổ hỏi do đang dừng hoặc ngủ.")
            # Nếu đang chạy nhưng vẫn đang ngủ (bị gọi sớm), lên lịch kiểm tra lại khi hết ngủ
            if self.is_running and time.time() < self.sleep_until_time:
                 self.schedule_next_reminder()
            return # Thoát khỏi hàm nếu không đủ điều kiện hiển thị

        self.play_notification_sound()

        # Tạo cửa sổ riêng (Toplevel) để hỏi
        ask_win = tk.Toplevel(self.root)
        ask_win.title("Bạn đang làm gì?")
        # ask_win.geometry("350x150") # Bỏ dòng này để dùng toàn màn hình
        ask_win.resizable(False, False) # Giữ nguyên không cho thay đổi kích thước thủ công

        # --- Yêu cầu: Chiếm trọn màn hình ---
        try:
            ask_win.attributes("-fullscreen", True)
            # ask_win.attributes("-topmost", True) # Có thể không cần thiết trong fullscreen
        except tk.TclError:
            # Nếu hệ điều hành không hỗ trợ fullscreen theo cách này
            print("Cảnh báo: Chế độ toàn màn hình Tkinter có thể không được hỗ trợ đầy đủ. Thử phóng to tối đa.")
            ask_win.state('zoomed') # Thử phóng to cửa sổ (có thể vẫn hiển thị thanh taskbar)
            ask_win.attributes("-topmost", True) # Đảm bảo ở trên cùng nếu không full màn hình

        # Đảm bảo cửa sổ luôn ở trên cùng và chiếm quyền tương tác
        ask_win.attributes("-topmost", True) # Đảm bảo luôn trên cùng
        ask_win.grab_set() # Chặn tương tác với cửa sổ khác cho đến khi cửa sổ này bị đóng

        # Tạo một Frame để chứa nội dung, giúp căn giữa nội dung trong cửa sổ fullscreen/maximized
        content_frame = tk.Frame(ask_win)
        # Sử dụng place để căn giữa Frame trong cửa sổ
        content_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Label câu hỏi
        label = tk.Label(content_frame, text=f"Bạn đã làm gì trong {self.reminder_interval_ms // (60 * 1000)} phút vừa qua?", font=("Helvetica", 16, "bold")) # Tăng kích thước font và in đậm cho dễ đọc trong fullscreen
        label.pack(pady=20)

        # Ô nhập liệu
        entry = tk.Entry(content_frame, width=60, font=("Helvetica", 14)) # Tăng width và font
        entry.pack(pady=10)
        entry.focus_set() # Đặt focus vào ô nhập liệu ngay lập tức

        # Hàm xử lý khi nhấn nút OK hoặc Enter
        def submit_answer():
            answer = entry.get().strip() # Lấy nội dung và loại bỏ khoảng trắng ở đầu/cuối

            # --- Yêu cầu: Không thể đóng nếu trống ---
            if not answer: # Kiểm tra xem ô nhập liệu có trống không
                messagebox.showwarning("Thông tin cần thiết", "Vui lòng nhập hoạt động của bạn trước khi xác nhận.", parent=ask_win)
                entry.focus_set() # Trả focus về ô nhập liệu để người dùng nhập
                return # Dừng hàm tại đây, không đóng cửa sổ

            # --- Nếu có nội dung, tiến hành lưu và đóng cửa sổ ---
            self.save_log(answer) # Lưu câu trả lời vào log Excel
            ask_win.destroy()      # Đóng cửa sổ hỏi
            self.schedule_next_reminder() # Lên lịch cho lần hỏi tiếp theo

        # Nút OK
        submit_button = tk.Button(content_frame, text="Xác nhận", command=submit_answer, width=15, font=("Helvetica", 14)) # Tăng width và font
        submit_button.pack(pady=20)

        # Xử lý khi người dùng nhấn phím Enter trong ô nhập liệu
        entry.bind("<Return>", lambda event: submit_answer())

        # --- Yêu cầu: Không thể đóng bằng nút 'X' ---
        # Ghi đè hành vi mặc định của nút đóng cửa sổ (nút 'X' ở góc)
        # Hàm này sẽ được gọi khi nhấn nút 'X'
        def prevent_closing():
             # Có thể hiển thị một thông báo hoặc chỉ đơn giản là không làm gì cả
             messagebox.showwarning("Không thể đóng", "Bạn cần nhập hoạt động của mình và nhấn 'Xác nhận' hoặc Enter.", parent=ask_win)
             entry.focus_set() # Trả focus về ô nhập liệu

        # Đăng ký hàm prevent_closing cho sự kiện WM_DELETE_WINDOW (nhấn nút X)
        ask_win.protocol("WM_DELETE_WINDOW", prevent_closing)

        # Không lên lịch nhắc nhở mới ngay lập tức ở đây,
        # mà đợi sau khi người dùng trả lời hoặc đóng cửa sổ hỏi.

    def go_to_sleep(self):
        """Tạm dừng nhắc nhở hoạt động trong một khoảng thời gian."""
        if self.countdown_id: # Nếu đang đếm ngược (đã ngủ rồi)
             messagebox.showinfo("Thông báo", "Đang trong chế độ ngủ rồi.", parent=self.root)
             return

        minutes = simpledialog.askinteger("Đi ngủ", "Bạn muốn nghỉ trong bao nhiêu phút?",
                                          minvalue=1, maxvalue=1440, parent=self.root) # Tối đa 1 ngày
        if minutes:
            # Hủy lịch trình nhắc nhở hoạt động tiếp theo (nếu có)
            if self.next_reminder_id:
                self.root.after_cancel(self.next_reminder_id)
                self.next_reminder_id = None
                print("Đã hủy lịch nhắc nhở hoạt động do đi ngủ.")

            self.sleep_until_time = time.time() + minutes * 60
            messagebox.showinfo("Thông báo", f"Sẽ không nhắc nhở hoạt động trong {minutes} phút tới.", parent=self.root)
            self.update_countdown() # Bắt đầu hiển thị đếm ngược trên status_label
            # schedule_next_reminder sẽ được gọi khi update_countdown hoàn thành
            # self.schedule_next_reminder() # Không cần gọi ở đây nữa

    def update_countdown(self):
        """Cập nhật hiển thị đếm ngược thời gian ngủ trên status_label."""
        # Hủy lịch cập nhật cũ nếu có
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None # Đặt lại về None sau khi hủy

        # Kiểm tra xem còn thời gian ngủ không
        remaining = int(self.sleep_until_time - time.time())

        if remaining > 0:
            mins, secs = divmod(remaining, 60)
            self.status_label.config(text=f"😴 Đang ngủ: còn {mins:02d}:{secs:02d}")
            # Lên lịch cập nhật lại sau 1 giây
            self.countdown_id = self.root.after(1000, self.update_countdown)
        else:
            # Hết giờ ngủ
            self.status_label.config(text="Đã thức dậy!")
            self.sleep_until_time = 0 # Đặt lại thời gian ngủ về 0
            self.countdown_id = None
            print("Hết giờ ngủ.")
            # Sau khi thức dậy, lên lịch nhắc nhở hoạt động tiếp theo
            self.schedule_next_reminder()


    def change_interval(self):
        """Thay đổi khoảng thời gian giữa các lần nhắc nhở hoạt động."""
        current_interval_min = self.reminder_interval_ms // (60 * 1000)
        new_minutes = simpledialog.askinteger("Cài đặt thời gian",
                                              f"Nhập khoảng thời gian nhắc nhở hoạt động (phút):\n(Hiện tại: {current_interval_min} phút)",
                                              minvalue=1, maxvalue=120, parent=self.root, # Giới hạn 2 tiếng
                                              initialvalue=current_interval_min)
        if new_minutes is not None and new_minutes != current_interval_min: # Kiểm tra new_minutes is not None vì simpledialog trả về None nếu nhấn Cancel
            self.reminder_interval_ms = new_minutes * 60 * 1000
            messagebox.showinfo("Thông báo", f"Đã cập nhật khoảng thời gian nhắc nhở hoạt động thành {new_minutes} phút.", parent=self.root)
            print(f"Đã đổi khoảng thời gian nhắc nhở hoạt động thành {new_minutes} phút.")
            # Lên lịch lại ngay với khoảng thời gian mới (nếu không đang ngủ)
            if time.time() >= self.sleep_until_time:
                 self.schedule_next_reminder()

    def stop_program(self):
        """Dừng ứng dụng hoàn toàn."""
        print("Yêu cầu dừng ứng dụng.")
        self.is_running = False # Đặt cờ dừng

        # Hủy tất cả các lịch trình .after() đang chờ
        if self.next_reminder_id:
            self.root.after_cancel(self.next_reminder_id)
            self.next_reminder_id = None
            print("Đã hủy lịch nhắc nhở hoạt động cuối cùng.")
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None
            print("Đã hủy lịch đếm ngược.")

        # --- Cần triển khai sau: Hủy tất cả báo thức nhiệm vụ ---
        # self.cancel_all_task_alarms()
        # ----------------------------------------------------

        print("Dừng ứng dụng.")
        self.root.destroy() # Đóng cửa sổ chính và kết thúc mainloop


# --- Khởi chạy ứng dụng ---
if __name__ == "__main__":
    root = tk.Tk()
    app = ReminderApp(root)
    # Đặt xử lý khi nhấn nút đóng cửa sổ (nút X) của cửa sổ chính
    root.protocol("WM_DELETE_WINDOW", app.stop_program)
    root.mainloop()

# file with big update about tasks todolist and alarm system
# 2025-04-24