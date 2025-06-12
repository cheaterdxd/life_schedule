import tkinter as tk
from tkinter import simpledialog, messagebox, PhotoImage
import time, threading
import os
import sys # Dùng để lấy đường dẫn nếu đóng gói thành exe

# Thư viện bên ngoài (cần cài đặt: pip install openpyxl playsound3)
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    messagebox.showerror("Lỗi thiếu thư viện", "Vui lòng cài đặt thư viện openpyxl: pip install openpyxl")
    exit()
try:
    from playsound3 import playsound
except ImportError:
    messagebox.showerror("Lỗi thiếu thư viện", "Vui lòng cài đặt thư viện playsound3: pip install playsound3")
    exit()

# --- Cấu hình ---
DEFAULT_INTERVAL_MINUTES = 15
EXCEL_FILE = "activity_log.xlsx"
SOUND_FILE = "clockbeep.wav" # Đặt file âm thanh cùng thư mục

# --- Hàm tiện ích tìm đường dẫn (hữu ích khi đóng gói bằng PyInstaller) ---
def resource_path(relative_path):
    """ Lấy đường dẫn tuyệt đối đến tài nguyên, hoạt động cho cả dev và PyInstaller """
    try:
        # PyInstaller tạo thư mục tạm _MEIPASS và lưu trữ tài nguyên ở đó
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- Lớp ứng dụng chính ---
class ReminderApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Trình Nhắc Nhở Hoạt Động")
        # self.root.geometry("300x150") # Kích thước ban đầu (tùy chỉnh)
        self.root.resizable(False, False) # Không cho thay đổi kích thước

        # --- Biến trạng thái ---
        self.reminder_interval_ms = DEFAULT_INTERVAL_MINUTES * 60 * 1000
        self.next_reminder_id = None # ID của lịch trình .after() tiếp theo
        self.sleep_until_time = 0    # Thời điểm kết thúc ngủ
        self.countdown_id = None     # ID của lịch trình countdown
        self.is_running = True       # Trạng thái chính của vòng lặp nhắc nhở

        # --- Âm thanh và Icon ---
        self.sound_path = resource_path(SOUND_FILE)
        # Thử tải icon (bỏ qua nếu không có file icon.png)
        try:
            icon_path = resource_path("icon.png")
            if os.path.exists(icon_path):
                 self.root.iconphoto(True, PhotoImage(file=icon_path))
        except Exception as e:
            print(f"Không thể tải icon: {e}")


        # --- Giao diện ---
        self.status_label = tk.Label(root_window, text="Đang chạy...", font=("Helvetica", 10), pady=10)
        self.status_label.pack(fill=tk.X, padx=10)

        # button_frame = tk.Frame(root_window)
        # button_frame.pack(pady=10)

        # self.sleep_button = tk.Button(button_frame, text="😴 Đi ngủ", command=self.go_to_sleep, width=10)
        # self.sleep_button.grid(row=0, column=0, padx=5)

        # self.settings_button = tk.Button(button_frame, text="⚙️ Cài đặt", command=self.change_interval, width=10)
        # self.settings_button.grid(row=0, column=1, padx=5)

        # self.stop_button = tk.Button(button_frame, text="⏹️ Dừng", command=self.stop_program, width=10)
        # self.stop_button.grid(row=0, column=2, padx=5)

        button_frame = tk.Frame(root_window)
        button_frame.pack(pady=10)

        # Nút Đi ngủ (Sleep Button) - Màu xanh nhẹ nhàng
        self.sleep_button = tk.Button(button_frame,
                                       text="😴 Đi ngủ",
                                       command=self.go_to_sleep,
                                       width=10,
                                       bg="lightblue",          # Màu nền xanh dương nhạt
                                       fg="black",              # Màu chữ đen
                                       activebackground="dodgerblue", # Màu nền khi nhấn (xanh sáng hơn)
                                       activeforeground="white",  # Màu chữ khi nhấn (trắng)
                                       relief="flat")           # Làm phẳng nút để màu nền hiển thị rõ
        self.sleep_button.grid(row=0, column=0, padx=5)

        # Nút Cài đặt (Settings Button) - Màu xám trung tính
        self.settings_button = tk.Button(button_frame,
                                         text="⚙️ Cài đặt",
                                         command=self.change_interval,
                                         width=10,
                                         bg="lightgray",        # Màu nền xám nhạt
                                         fg="black",            # Màu chữ đen
                                         activebackground="gray", # Màu nền khi nhấn (xám đậm hơn)
                                         activeforeground="white", # Màu chữ khi nhấn (trắng)
                                         relief="flat")         # Làm phẳng nút
        self.settings_button.grid(row=0, column=1, padx=5)

        # Nút Dừng (Stop Button) - Màu đỏ báo hiệu
        self.stop_button = tk.Button(button_frame,
                                     text="⏹️ Dừng",
                                     command=self.stop_program,
                                     width=10,
                                     bg="red",              # Màu nền đỏ
                                     fg="white",            # Màu chữ trắng
                                     activebackground="darkred", # Màu nền khi nhấn (đỏ đậm)
                                     activeforeground="white", # Màu chữ khi nhấn (trắng)
                                     relief="flat")         # Làm phẳng nút
        self.stop_button.grid(row=0, column=2, padx=5)


        # Đảm bảo file Excel tồn tại với tiêu đề
        self.ensure_excel_file()

        # Bắt đầu vòng lặp nhắc nhở đầu tiên
        self.schedule_next_reminder()

    def ensure_excel_file(self):
        """Kiểm tra và tạo file Excel nếu chưa có."""
        excel_path = resource_path(EXCEL_FILE)
        if not os.path.exists(excel_path):
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(["Thời gian", "Hoạt động"])
                wb.save(excel_path)
                print(f"Đã tạo file log: {excel_path}")
            except Exception as e:
                messagebox.showerror("Lỗi File Excel", f"Không thể tạo file {EXCEL_FILE}:\n{e}")
                self.stop_program() # Dừng nếu không tạo được file log

    def save_log(self, answer):
        """Lưu câu trả lời vào file Excel."""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        excel_path = resource_path(EXCEL_FILE)
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            ws.append([timestamp, answer])
            wb.save(excel_path)
        except Exception as e:
             messagebox.showwarning("Lỗi Lưu Log", f"Không thể ghi vào file {EXCEL_FILE}:\n{e}\n\nLog sẽ chỉ được lưu tạm vào bộ nhớ.")
             # Có thể thêm cơ chế lưu tạm vào biến hoặc file text dự phòng ở đây

    def play_notification_sound(self):
        """Phát âm thanh thông báo."""
        try:
            # Chạy playsound trong thread riêng để tránh treo GUI nếu file lớn/lỗi
            threading.Thread(target=playsound, args=(self.sound_path,), daemon=True).start()
        except Exception as e:
            print(f"Lỗi khi phát âm thanh '{self.sound_path}': {e}")
            # Không cần messagebox ở đây để tránh làm phiền liên tục

    def schedule_next_reminder(self):
        """Lên lịch cho lần nhắc nhở tiếp theo."""
        # Hủy lịch trình cũ nếu có
        if self.next_reminder_id:
            self.root.after_cancel(self.next_reminder_id)

        if self.is_running and time.time() >= self.sleep_until_time:
            print(f"Lên lịch nhắc nhở sau {self.reminder_interval_ms // 1000} giây.")
            self.next_reminder_id = self.root.after(self.reminder_interval_ms, self.ask_question)
            self.status_label.config(text="Đang chạy...")
        elif self.is_running and time.time() < self.sleep_until_time:
            # Nếu đang ngủ, kiểm tra lại khi hết giờ ngủ
             wake_up_delay = int((self.sleep_until_time - time.time()) * 1000)
             print(f"Đang ngủ, kiểm tra lại sau {wake_up_delay / 1000:.0f} giây.")
             self.next_reminder_id = self.root.after(wake_up_delay + 500, self.schedule_next_reminder) # Thêm 0.5s để chắc chắn hết ngủ

    def ask_question(self):
        """Hỏi người dùng và lên lịch lại."""
        if not self.is_running or time.time() < self.sleep_until_time:
            # Trường hợp bị gọi nhầm khi đang dừng hoặc ngủ
            self.schedule_next_reminder()
            return

        self.play_notification_sound()

        # Tạo cửa sổ riêng thay vì simpledialog để kiểm soát tốt hơn
        ask_win = tk.Toplevel(self.root)
        ask_win.title("Bạn đang làm gì?")
        ask_win.geometry("350x150")
        ask_win.resizable(False, False)
        # Đưa lên trên cùng
        ask_win.attributes("-topmost", True)
        # Focus vào cửa sổ này
        ask_win.grab_set()

        label = tk.Label(ask_win, text=f"Bạn đã làm gì trong {self.reminder_interval_ms // (60 * 1000)} phút vừa qua?")
        label.pack(pady=10)

        entry = tk.Entry(ask_win, width=40)
        entry.pack(pady=5)
        entry.focus_set() # Focus vào ô nhập liệu

        def submit_answer():
            answer = entry.get()
            if answer:
                self.save_log(answer)
            else:
                self.save_log("Không trả lời") # Ghi lại nếu không nhập gì
            ask_win.destroy() # Đóng cửa sổ hỏi
            self.schedule_next_reminder() # Lên lịch cho lần hỏi tiếp theo

        submit_button = tk.Button(ask_win, text="OK", command=submit_answer, width=10)
        submit_button.pack(pady=10)

        # Xử lý khi nhấn Enter
        entry.bind("<Return>", lambda event: submit_answer())
        # Xử lý khi đóng cửa sổ bằng nút 'X'
        ask_win.protocol("WM_DELETE_WINDOW", lambda: (self.save_log("Đóng cửa sổ"), ask_win.destroy(), self.schedule_next_reminder()))

        # Không lên lịch nhắc nhở mới ngay lập tức ở đây,
        # mà đợi sau khi người dùng trả lời hoặc đóng cửa sổ hỏi.

    def go_to_sleep(self):
        """Tạm dừng nhắc nhở trong một khoảng thời gian."""
        if self.countdown_id: # Nếu đang đếm ngược (đã ngủ rồi)
             messagebox.showinfo("Thông báo", "Đang trong chế độ ngủ rồi.", parent=self.root)
             return

        minutes = simpledialog.askinteger("Đi ngủ", "Bạn muốn nghỉ trong bao nhiêu phút?",
                                          minvalue=1, maxvalue=1440, parent=self.root)
        if minutes:
            # Hủy lịch trình nhắc nhở tiếp theo
            if self.next_reminder_id:
                self.root.after_cancel(self.next_reminder_id)
                self.next_reminder_id = None
                print("Đã hủy lịch nhắc nhở do đi ngủ.")

            self.sleep_until_time = time.time() + minutes * 60
            messagebox.showinfo("Thông báo", f"Sẽ không nhắc nhở trong {minutes} phút tới.", parent=self.root)
            self.update_countdown() # Bắt đầu hiển thị đếm ngược
            # Lên lịch để kiểm tra lại khi hết giờ ngủ
            self.schedule_next_reminder()

    def update_countdown(self):
        """Cập nhật hiển thị đếm ngược thời gian ngủ."""
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id) # Hủy lịch cũ

        if time.time() < self.sleep_until_time:
            remaining = int(self.sleep_until_time - time.time())
            mins, secs = divmod(remaining, 60)
            self.status_label.config(text=f"😴 Đang ngủ: còn {mins:02d}:{secs:02d}")
            self.countdown_id = self.root.after(1000, self.update_countdown) # Lên lịch cập nhật sau 1s
        else:
            # Hết giờ ngủ
            self.status_label.config(text="Đã thức dậy!")
            self.sleep_until_time = 0
            self.countdown_id = None
            print("Hết giờ ngủ.")
            # Không cần gọi schedule_next_reminder ở đây vì nó đã được gọi trong go_to_sleep
            # để kiểm tra lại khi hết giờ ngủ.

    def change_interval(self):
        """Thay đổi khoảng thời gian giữa các lần nhắc nhở."""
        current_interval_min = self.reminder_interval_ms // (60 * 1000)
        new_minutes = simpledialog.askinteger("Cài đặt thời gian",
                                              f"Nhập khoảng thời gian nhắc nhở (phút):\n(Hiện tại: {current_interval_min} phút)",
                                              minvalue=1, maxvalue=120, parent=self.root, # Giới hạn 2 tiếng
                                              initialvalue=current_interval_min)
        if new_minutes and new_minutes != current_interval_min:
            self.reminder_interval_ms = new_minutes * 60 * 1000
            messagebox.showinfo("Thông báo", f"Đã cập nhật khoảng thời gian nhắc nhở thành {new_minutes} phút.", parent=self.root)
            print(f"Đã đổi khoảng thời gian thành {new_minutes} phút.")
            # Lên lịch lại ngay với khoảng thời gian mới (nếu không đang ngủ)
            if time.time() >= self.sleep_until_time:
                 self.schedule_next_reminder()

    def stop_program(self):
        """Dừng ứng dụng hoàn toàn."""
        self.is_running = False
        if self.next_reminder_id:
            self.root.after_cancel(self.next_reminder_id)
            print("Đã hủy lịch nhắc nhở cuối cùng.")
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            print("Đã hủy lịch đếm ngược.")

        print("Dừng ứng dụng.")
        self.root.destroy() # Đóng cửa sổ chính

# --- Khởi chạy ứng dụng ---
if __name__ == "__main__":
    root = tk.Tk()
    app = ReminderApp(root)
    # Đặt xử lý khi nhấn nút đóng cửa sổ (nút X)
    root.protocol("WM_DELETE_WINDOW", app.stop_program)
    root.mainloop()